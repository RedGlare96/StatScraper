import random
import time
import logging
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import undetected_chromedriver as uc
from selenium_stealth import stealth
import os


class ScraperBase:

    logger = logging.getLogger('ScraperCore')

    def __init__(self, url, outputpath, single_process=False, use_proxy=False, use_stealth=False):
        self.url = url
        self.single_process = single_process
        self.use_proxy = use_proxy
        self.use_stealth = use_stealth
        self.driver = None
        if not use_proxy:
            self.cookie_file = 'chrome-data'
        else:
            self.cookie_file = 'zenmate-cookies'
        self.bin_dir = 'chrome-bin/chrome.exe'
        self.outputpath = outputpath
        self.output = dict()

    def get_data(self):
        self.logger.error('Base method called')

    def run_browser(self):
        self.logger.info('Init browser')
        options = uc.ChromeOptions()
        options.user_data_dir = self.cookie_file
        options.binary_location = self.bin_dir
        options.add_argument('--no-first-run --no-service-autorun --password-store=basic')
        options.add_argument('--window-size={}'.format('1920,1080'))
        options.add_argument('--disk-cache-size=1073741824')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-gpu')
        options.add_argument('--dns-prefetch-disable')
        options.add_argument('--hide-scrollbars')
        options.add_argument("--disable-infobars")
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-browser-side-navigation')
        options.add_argument('--log-level=0')
        if self.single_process:
            options.add_argument('--single-process')
        options.add_argument('--ignore-certificate-errors')
        options.add_argument("--disable-plugins-discovery")
        options.add_argument("--start-maximized")
        if self.use_proxy:
            options.add_argument(
                f"--load-extension={os.path.join(os.path.dirname(os.path.abspath(__file__)), 'plugin', 'zenmate')}")
        self.driver = uc.Chrome(headless=False, options=options, version_main=102)
        if self.use_stealth:
            stealth(self.driver,
                    languages=["en-US", "en"],
                    vendor="Google Inc.",
                    platform="Win32",
                    webgl_vendor="Intel Inc.",
                    renderer="Intel Iris OpenGL Engine",
                    fix_hairline=True,
                    )
        self.get_data()

    def save_into_file(self):
        '''
        Save data into file
        '''
        logger = logging.getLogger(__name__ + 'SaveFile')
        if os.path.isfile(self.outputpath):
            wb = load_workbook(self.outputpath)
            ws = wb.active
            next_row = len(ws['A']) + 1
            for index, value in enumerate(self.output.values()):
                ws.cell(row=next_row, column=index + 1, value=value)
            logger.debug('Saving entry to: {0}'.format(self.outputpath))
            wb.save(self.outputpath)
        else:
            # Existing file with previous data not found. Assumed to be first time scraping
            logger.debug('New file created')
            wb = Workbook()
            ws = wb.active
            for index, (key, value) in enumerate(self.output.items()):
                # Filling column titles
                ws.cell(row=1, column=index + 1, value=key)
                # Filling first row
                ws.cell(row=2, column=index + 1, value=value)
            logger.debug('Saving entry to: {0}'.format(self.outputpath))
            wb.save(self.outputpath)


class TrustpilotScraper(ScraperBase):

    def __int__(self, url, outputpath):
        super(TrustpilotScraper, self).__init__(url, outputpath, single_process=True)

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(2)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            logo_div = prod_soup.find('div', {'class': 'profile-image_imageWrapper__kDdWe'})
            logo_image = logo_div.find_all('source')[1]['srcset'].split(',')[-1].replace('2x', '').strip()
            self.logger.debug('Logo image: {}'.format(logo_image))
            self.output['Logo Image'] = logo_image
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Logo image'] = 'N/A'
        main_div = prod_soup.find('div', {'id': 'business-unit-title'})
        try:
            name = main_div.find('h1').text.replace('Reviews', '').strip()
            self.logger.debug('Name: {}'.format(name))
            self.output['Name'] = name
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Name'] = 'N/A'
        try:
            tot_rew = main_div.find('span', {'class': 'typography_typography__QgicV typography'
                                                                         '_bodysmall__irytL typography_color-gray-7__9Ut3K'
                                                                         ' typography_weight-regular__TWEnf typography'
                                                                         '_fontstyle-normal__kHyN3'
                                                      ' styles_text__W4hWi'}).text.split()[0].strip()
            self.logger.debug('Total Reviews: {}'.format(tot_rew))
            self.output['Total Reviews'] = tot_rew
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Total Reviews'] = 'N/A'
        try:
            stars = main_div.find('div', {'class': 'styles_container__OaEK8'}).find('p').text
            self.logger.debug('Main star rating: {}'.format(stars))
            self.output['Main star rating'] = stars
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Main star rating'] = 'N/A'
        star_range = None
        star_image = None
        try:
            rating_box = main_div.find('div', {'class': 'star-rating_starRating__4rrcf star-rating_medium__iN6Ty'})
            star_range = rating_box.find('img')['alt']
            self.logger.debug('Star range: {}'.format(rating_box.find('img')['alt']))
            self.output['Star range'] = star_range
            star_image = rating_box.find('img')['src']
            self.logger.debug('Star Image: {}'.format(star_image))
            self.output['Star Image'] = star_image
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            if star_range is None:
                self.output['star_range'] = 'N/A'
            if star_image is None:
                self.output['star_image'] = 'N/A'
        try:
            verified = 'Yes' if main_div.find('button', {'class': 'styles_verificationLabel__kukuk'}) is not None else 'No'
            self.logger.debug('Verified: {}'.format(verified))
            self.output['Verified'] = verified
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Main star rating'] = 'N/A'
        self.save_into_file()


class FacebookScraper(ScraperBase):

    def __init__(self, url, outputpath):
        super(FacebookScraper, self).__init__(url, outputpath, single_process=True)

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(2)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            name = prod_soup.find('h1').text
            self.logger.debug('Name: {}'.format(name))
            self.output['Name'] = name
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Name'] = 'N/A'
        try:
            likes = prod_soup.find('span', {'class': 'd2edcug0 hpfvmrgz qv66sw1b c1et5uql lr9zc1uh jq4qci2q'
                                                     ' a3bd9o3v b1v8xokw oo9gr5id'}).text
            self.logger.debug('Likes: {}'.format(likes))
            self.output['Likes'] = likes
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Likes'] = 'N/A'
        follows1 = None
        follows2 = None
        try:
            follows1 = 'Follows: {}'.format(prod_soup.find('span', {'class': 'd2edcug0 hpfvmrgz qv66sw1b c1et5uql'
                                                                             ' lr9zc1uh jq4qci2q a3bd9o3v b1v8xokw'
                                                                             ' oo9gr5id'}).text)
            self.logger.debug('Follows 1: {}'.format(follows1))
            self.output['Follows1'] = follows1
            follows2 = prod_soup.find('a', {'class': 'oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv'
                                                     ' nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab'
                                                     ' hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h'
                                                     ' esuyzwwr f1sip0of lzcic4wl gpro0wi8 m9osqain lrazzd5p'}).text
            self.logger.debug('Follows 2: {}'.format(follows2))
            self.output['Follows1'] = follows2
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            if follows1 is None:
                self.output['Follows1'] = 'N/A'
            if follows2 is None:
                self.output['Follows2'] = 'N/A'
        self.save_into_file()


class TwitterScraper(ScraperBase):

    def __init__(self, url, outputpath):
        super(TwitterScraper, self).__init__(url, outputpath, single_process=True)

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(2)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            name = prod_soup.find_all('span', {'class': 'css-901oao css-16my406 r-poiln3 r-bcqeeo r-qvutc0'})[6].text
            self.logger.debug('Name: {}'.format(name))
            self.output['Name'] = name
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Name'] = 'N/A'
        try:
            followers = prod_soup.find('span', string='Followers').find_previous().text
            self.logger.debug('Followers: {}'.format(followers))
            self.output['Followers'] = followers
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Followers'] = 'N/A'
        self.save_into_file()


class LinkedinScraper(ScraperBase):

    def __init__(self, url, outputpath):
        super(LinkedinScraper, self).__init__(url, outputpath, use_stealth=True)

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(5)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            name = prod_soup.find('h1', {'class': 'top-card-layout__title font-sans text-lg papabear:text-xl'
                                                  ' font-bold leading-open text-color-text mb-0'}).text.strip()
            self.logger.debug('Name: {}'.format(name))
            self.output['Name'] = name
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Name'] = 'N/A'
        try:
            followers = prod_soup.find('h3', {'class': 'top-card-layout__first-subline font-sans text-md leading-open'
                                                        ' text-color-text-low-emphasis'}).text.strip().split()[2]
            self.logger.debug('Followers: {}'.format(followers))
            self.output['Followers'] = followers
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Followers'] = 'N/A'
        self.save_into_file()


class YoutubeScraper(ScraperBase):

    def __init__(self, url, outputpath):
        super(YoutubeScraper, self).__init__(url, outputpath, single_process=True)

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(2)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            name = prod_soup.find('yt-formatted-string', {'id': 'text', 'class': 'style-scope ytd-channel-name'}).text
            self.logger.debug('Name: {}'.format(name))
            self.output['Name'] = name
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Name'] = 'N/A'
        try:
            subscribers = prod_soup.find('yt-formatted-string', {'id': 'subscriber-count'}).text
            self.logger.debug('Subscribers: {}'.format(subscribers))
            self.output['Subscribers'] = subscribers
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Subscribers'] = 'N/A'
        self.save_into_file()


class TiktokScraper(ScraperBase):

    def __init__(self, url, outputpath):
        super(TiktokScraper, self).__init__(url, outputpath, use_proxy=True)

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(5)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            name = prod_soup.find('h2', {'class': 'tiktok-b7g450-H2ShareTitle ekmpd5l5'}).text
            self.logger.debug('Name: {}'.format(name))
            self.output['Name'] = name
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Name'] = 'N/A'
        try:
            followers = prod_soup.find_all('div', {'class': 'tiktok-xeexlu-DivNumber e1457k4r1'})[1].find('strong').text
            self.logger.debug('Followers: {}'.format(followers))
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Followers'] = 'N/A'
        try:
            likes = prod_soup.find_all('div', {'class': 'tiktok-xeexlu-DivNumber e1457k4r1'})[2].find('strong').text
            self.logger.debug('Likes: {}'.format(likes))
            self.output['Likes'] = likes
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Likes'] = 'N/A'
        self.save_into_file()


class InstagramScraper(ScraperBase):

    def __init__(self, url, outputpath):
        super(InstagramScraper, self).__init__(url, outputpath, use_stealth=True)

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(5)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            title = prod_soup.find('h2').text
            self.logger.debug('Title: {}'.format(title))
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Title'] = 'N/A'
        try:
            followers = prod_soup.find_all('li', {'class': 'Y8-fY'})[1].find('span')['title']
            self.logger.debug('Followers: {}'.format(followers))
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Followers'] = 'N/A'
        self.save_into_file()

