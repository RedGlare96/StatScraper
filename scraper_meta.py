from PIL import Image
import time
import logging
from io import BytesIO
import openpyxl.drawing.image
from bs4 import BeautifulSoup
from requests import get
from openpyxl import Workbook, load_workbook, worksheet
from openpyxl import drawing as dr
import os


class ScraperBase:

    logger = logging.getLogger('ScraperCore')

    def __init__(self, url, outputpath, driver):
        self.url = url
        self.driver = driver
        self.outputpath = outputpath
        self.output = dict()

    @staticmethod
    def get_keyword():
        return 'base'

    def get_data(self):
        raise Exception('Base method called')

    def run_browser(self):
        self.logger.info('Entering in new tab')
        self.logger.debug('Opening new tab')
        self.driver.switch_to.new_window()
        self.logger.debug('Switching handle')
        self.driver.switch_to.window(self.driver.window_handles[0])
        self.logger.debug('Closing old tab')
        self.driver.close()
        self.logger.info('Switching handle 2')
        self.driver.switch_to.window(self.driver.window_handles[0])
        self.get_data()

    def process_image(self, image_url):
        ext = image_url.split('.')[-1]
        self.logger.debug('Image ext: {}'.format(ext))
        img_path = 'saved-image.{}'.format(ext)
        r = get(image_url.replace('imagesrc:', ''))
        image = Image.open((BytesIO(r.content)))
        image = image.resize((140, 90))
        image.save(img_path)
        return img_path

    def save_into_file(self):
        '''
        Save data into file
        '''
        logger = logging.getLogger(__name__ + '.SaveFile')
        if os.path.isfile(self.outputpath):
            wb = load_workbook(self.outputpath)
            if self.get_keyword() not in wb.sheetnames:
                logger.debug('Creating new sheet')
                ws = wb.create_sheet(self.get_keyword())
                for index, (key, value) in enumerate(self.output.items()):
                    # Filling column titles
                    ws.cell(row=1, column=index + 1, value=key)
                    # Filling first row
                    if 'imagesrc:' in value:
                        try:
                            logger.debug('Processing image')
                            logger.debug('Changing cell dimensions')
                            ws.row_dimensions[2].height = 70
                            ws.column_dimensions[chr(ord('@') + index + 1)].width = 20
                            img = openpyxl.drawing.image.Image(self.process_image(value))
                            cell_name = '{0}{1}'.format(chr(ord('@') + index + 1), 2)
                            logger.debug('Using cell: {}'.format(cell_name))
                            ws.add_image(img, cell_name)
                        except Exception as exc:
                            logger.error('Error with image processing. Entering default value')
                            logger.debug('Details: {}'.format(str(exc)))
                            ws.cell(row=2, column=index + 1, value=value.replace('imagesrc:', ''))
                    else:
                        ws.cell(row=2, column=index + 1, value=value)
            else:
                ws = wb[self.get_keyword()]
                next_row = len(ws['A']) + 1
                for index, value in enumerate(self.output.values()):
                    if 'imagesrc:' in value:
                        try:
                            logger.debug('Processing image')
                            logger.debug('Changing cell dimensions')
                            ws.row_dimensions[next_row].height = 70
                            ws.column_dimensions[chr(ord('@') + index + 1)].width = 20
                            img = openpyxl.drawing.image.Image(self.process_image(value))
                            cell_name = '{0}{1}'.format(chr(ord('@') + index + 1), next_row)
                            logger.debug('Using cell: {}'.format(cell_name))
                            ws.add_image(img, cell_name)
                        except Exception as exc:
                            logger.error('Error with image processing. Entering default value')
                            logger.debug('Details: {}'.format(str(exc)))
                            ws.cell(row=next_row, column=index + 1, value=value.replace('imagesrc:', ''))
                    else:
                        ws.cell(row=next_row, column=index + 1, value=value)
        else:
            # Existing file with previous data not found. Assumed to be first time scraping
            logger.debug('New file created')
            wb = Workbook()
            logger.debug('Creating new sheet')
            ws = wb.active
            ws.title = self.get_keyword()
            for index, (key, value) in enumerate(self.output.items()):
                # Filling column titles
                ws.cell(row=1, column=index + 1, value=key)
                # Filling first row
                if 'imagesrc:' in value:
                    try:
                        logger.debug('Processing image')
                        logger.debug('Changing cell dimensions')
                        ws.row_dimensions[2].height = 70
                        ws.column_dimensions[chr(ord('@') + index + 1)].width = 20
                        img = openpyxl.drawing.image.Image(self.process_image(value))
                        cell_name = '{0}{1}'.format(chr(ord('@') + index + 1), 2)
                        logger.debug('Using cell: {}'.format(cell_name))
                        ws.add_image(img, cell_name)
                    except Exception as exc:
                        logger.error('Error with image processing. Entering default value')
                        logger.debug('Details: {}'.format(str(exc)))
                        ws.cell(row=2, column=index + 1, value=value.replace('imagesrc:', ''))
                else:
                    ws.cell(row=2, column=index + 1, value=value)
        logger.debug('Saving entry to: {0}'.format(self.outputpath))
        wb.save(self.outputpath)


class TrustpilotScraper(ScraperBase):

    def __int__(self, url, outputpath, driver):
        super(TrustpilotScraper, self).__init__(url, outputpath, driver)

    @staticmethod
    def get_keyword():
        return 'trustpilot'

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(2)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            logo_div = prod_soup.find('div', {'class': 'profile-image_imageWrapper__kDdWe'})
            logo_image = 'imagesrc:' + logo_div.find_all('source')[1]['srcset'].split(',')[-1].split()[0].strip()
            self.logger.debug('Logo image: {}'.format(logo_image))
            self.output['Logo Image'] = logo_image
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Logo image'] = 'N/A'
        main_div = prod_soup.find('div', {'id': 'business-unit-title'})
        try:
            name = main_div.find('h1').text.replace('Reviews', '').strip()
            self.logger.debug('Name: {}'.format(name))
            self.output['Name'] = name
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Name'] = 'N/A'
        try:
            tot_rew = main_div.find('span', {'class': 'typography_typography__QgicV typography'
                                                                         '_bodysmall__irytL typography_color-gray-7__9Ut3K'
                                                                         ' typography_weight-regular__TWEnf typography'
                                                                         '_fontstyle-normal__kHyN3'
                                                      ' styles_text__W4hWi'}).text.split()[0].strip()
            self.logger.debug('Total Reviews: {}'.format(tot_rew))
            self.output['Total Reviews'] = tot_rew
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Total Reviews'] = 'N/A'
        try:
            stars = main_div.find('div', {'class': 'styles_container__OaEK8'}).find('p').text
            self.logger.debug('Main star rating: {}'.format(stars))
            self.output['Main star rating'] = stars
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Main star rating'] = 'N/A'
        star_range = None
        star_image = None
        try:
            rating_box = main_div.find('div', {'class': 'star-rating_starRating__4rrcf star-rating_medium__iN6Ty'})
            star_range = rating_box.find('img')['alt']
            self.logger.debug('Star range: {}'.format(rating_box.find('img')['alt']))
            self.output['Star range'] = star_range
            star_image = 'imagesrc:' + rating_box.find('img')['src']
            self.logger.debug('Star Image: {}'.format(star_image))
            self.output['Star Image'] = star_image
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            if star_range is None:
                self.output['star_range'] = 'N/A'
            if star_image is None:
                self.output['star_image'] = 'N/A'
        try:
            verified = 'Yes' if main_div.find('button', {'class': 'styles_verificationLabel__kukuk'}) is not None else 'No'
            self.logger.debug('Verified: {}'.format(verified))
            self.output['Verified'] = verified
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Main star rating'] = 'N/A'
        self.save_into_file()


class FacebookScraper(ScraperBase):

    def __init__(self, url, outputpath, driver):
        super(FacebookScraper, self).__init__(url, outputpath, driver)

    @staticmethod
    def get_keyword():
        return 'facebook'

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(2)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            name = prod_soup.find('h1').text
            self.logger.debug('Name: {}'.format(name))
            self.output['Name'] = name
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Name'] = 'N/A'
        try:
            likes = prod_soup.find('span', {'class': 'd2edcug0 hpfvmrgz qv66sw1b c1et5uql lr9zc1uh jq4qci2q'
                                                     ' a3bd9o3v b1v8xokw oo9gr5id'}).text
            self.logger.debug('Likes: {}'.format(likes))
            self.output['Likes'] = likes
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Likes'] = 'N/A'
        follows1 = None
        follows2 = None
        try:
            follows1 = 'Follows: {}'.format(prod_soup.find_all('span', {'class': 'd2edcug0 hpfvmrgz qv66sw1b c1et5uql'
                                                                             ' lr9zc1uh jq4qci2q a3bd9o3v b1v8xokw'
                                                                             ' oo9gr5id'})[1].text)
            self.logger.debug('Follows 1: {}'.format(follows1))
            self.output['Follows1'] = follows1
            follows2 = prod_soup.find('a', {'class': 'oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv'
                                                     ' nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab'
                                                     ' hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h'
                                                     ' esuyzwwr f1sip0of lzcic4wl gpro0wi8 m9osqain lrazzd5p'}).text
            self.logger.debug('Follows 2: {}'.format(follows2))
            self.output['Follows1'] = follows2
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            if follows1 is None:
                self.output['Follows1'] = 'N/A'
            if follows2 is None:
                self.output['Follows2'] = 'N/A'
        self.save_into_file()


class TwitterScraper(ScraperBase):

    def __init__(self, url, outputpath, driver):
        super(TwitterScraper, self).__init__(url, outputpath, driver)

    @staticmethod
    def get_keyword():
        return 'twitter'

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(5)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            name = prod_soup.find_all('span', {'class': 'css-901oao css-16my406 r-poiln3 r-bcqeeo r-qvutc0'})[10].text
            self.logger.debug('Name: {}'.format(name))
            self.output['Name'] = name
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Name'] = 'N/A'
        try:
            followers = prod_soup.find('span', string='Followers').find_previous().text
            self.logger.debug('Followers: {}'.format(followers))
            self.output['Followers'] = followers
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Followers'] = 'N/A'
        self.save_into_file()


class LinkedinScraper(ScraperBase):

    def __init__(self, url, outputpath, driver):
        super(LinkedinScraper, self).__init__(url, outputpath, driver)

    @staticmethod
    def get_keyword():
        return 'linkedin'

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(5)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            name = prod_soup.find('h1', {'class': 'top-card-layout__title font-sans text-lg papabear:text-xl'
                                                  ' font-bold leading-open text-color-text mb-0'}).text.strip()
            self.logger.debug('Name: {}'.format(name))
            self.output['Name'] = name
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Name'] = 'N/A'
        try:
            followers = prod_soup.find('h3', {'class': 'top-card-layout__first-subline font-sans text-md leading-open'
                                                        ' text-color-text-low-emphasis'}).text.strip().split()[2]
            self.logger.debug('Followers: {}'.format(followers))
            self.output['Followers'] = followers
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Followers'] = 'N/A'
        self.save_into_file()


class YoutubeScraper(ScraperBase):

    def __init__(self, url, outputpath, driver):
        super(YoutubeScraper, self).__init__(url, outputpath, driver)

    @staticmethod
    def get_keyword():
        return 'youtube'

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(2)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            name = prod_soup.find('yt-formatted-string', {'id': 'text', 'class': 'style-scope ytd-channel-name'}).text
            self.logger.debug('Name: {}'.format(name))
            self.output['Name'] = name
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Name'] = 'N/A'
        try:
            subscribers = prod_soup.find('yt-formatted-string', {'id': 'subscriber-count'}).text
            self.logger.debug('Subscribers: {}'.format(subscribers))
            self.output['Subscribers'] = subscribers
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Subscribers'] = 'N/A'
        self.save_into_file()


class TiktokScraper(ScraperBase):

    def __init__(self, url, outputpath, driver):
        super(TiktokScraper, self).__init__(url, outputpath, driver)

    @staticmethod
    def get_keyword():
        return 'tiktok'

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(5)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            name = prod_soup.find('h2', {'class': 'tiktok-b7g450-H2ShareTitle ekmpd5l5'}).text
            self.logger.debug('Name: {}'.format(name))
            self.output['Name'] = name
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Name'] = 'N/A'
        try:
            followers = prod_soup.find_all('div', {'class': 'tiktok-xeexlu-DivNumber e1457k4r1'})[1].find('strong').text
            self.logger.debug('Followers: {}'.format(followers))
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Followers'] = 'N/A'
        try:
            likes = prod_soup.find_all('div', {'class': 'tiktok-xeexlu-DivNumber e1457k4r1'})[2].find('strong').text
            self.logger.debug('Likes: {}'.format(likes))
            self.output['Likes'] = likes
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Likes'] = 'N/A'
        self.save_into_file()


class InstagramScraper(ScraperBase):

    def __init__(self, url, outputpath, driver):
        super(InstagramScraper, self).__init__(url, outputpath, driver)

    @staticmethod
    def get_keyword():
        return 'instagram'

    def get_data(self):
        self.logger.info('Starting scraping process')
        self.driver.get(self.url)
        time.sleep(5)
        prod_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        try:
            title = prod_soup.find('h2').text
            self.logger.debug('Title: {}'.format(title))
            self.output['Title'] = title
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Title'] = 'N/A'
        try:
            followers = prod_soup.find_all('li', {'class': 'Y8-fY'})[1].find('span')['title']
            self.logger.debug('Followers: {}'.format(followers))
        except Exception as exc:
            self.logger.error('Could not get data')
            self.logger.debug('Details: {}'.format(str(exc)))
            self.output['Followers'] = 'N/A'
        self.save_into_file()

