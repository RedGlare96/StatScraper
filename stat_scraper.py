import random
from sys import stdout
import logging
from datetime import datetime
import time
from openpyxl import load_workbook
import undetected_chromedriver as uc
from selenium_stealth import stealth
import os
from os import path
from scraper_meta import ScraperBase

bin_dir = 'chrome-bin/chrome.exe'
version_number = 102
use_proxy = True
debug_mode = False
link_list = 'link_list.xlsx'
out_path = 'test1.xlsx'


def check_create_dir(dirname):
    '''
    Checks if directory exists and if it doesn't creates a new directory
    :param dirname: Path to directory
    '''
    if not path.exists(dirname):
        if '/' in dirname:
            os.makedirs(dirname)
        else:
            os.mkdir(dirname)


def read_links(inputfile):
    wb = load_workbook(inputfile)
    ws = wb.active
    ret = [x.value for x in ws['A']]
    return ret


if __name__ == '__main__':
    # Init logging
    rootLogger = logging.getLogger()
    consoleHandler = logging.StreamHandler(stdout)
    check_create_dir('logs')
    log_timestamp = datetime.now()
    fileHandler = logging.FileHandler(
        path.join('logs', 'CfdScraper{0}.log'.format(log_timestamp.strftime('%d-%m-%y-%H-%M-%S'))))
    fileHandler.setFormatter(logging.Formatter('%(asctime)s:-[%(name)s] - %(levelname)s - %(message)s'))
    rootLogger.addHandler(consoleHandler)
    rootLogger.addHandler(fileHandler)
    rootLogger.setLevel(logging.DEBUG)
    logging.getLogger('seleniumwire.handler').setLevel(logging.ERROR)
    logging.getLogger('selenium.webdriver.remote.remote_connection').setLevel(logging.ERROR)
    logging.getLogger('seleniumwire.server').setLevel(logging.ERROR)
    logging.getLogger('hpack.hpack').setLevel(logging.ERROR)
    logging.getLogger('hpack.table').setLevel(logging.ERROR)
    logging.getLogger('seleniumwire.storage').setLevel(logging.ERROR)
    if debug_mode:
        consoleHandler.setLevel(logging.DEBUG)
    else:
        consoleHandler.setLevel(logging.INFO)
    fileHandler.setLevel(logging.DEBUG)
    consoleHandler.setFormatter(logging.Formatter('[%(name)s] - %(levelname)s - %(message)s'))

    if use_proxy:
        cookie_file = 'zenmate-cookies'
    else:
        cookie_file = 'chrome-data'

    rootLogger.info('Initiating driver')
    options = uc.ChromeOptions()
    options.user_data_dir = cookie_file
    options.binary_location = bin_dir
    options.add_argument('--no-first-run --no-service-autorun --password-store=basic')
    options.add_argument('--window-size={}'.format('1920,1080'))
    options.add_argument('--disk-cache-size=1073741824')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-gpu')
    options.add_argument('--dns-prefetch-disable')
    options.add_argument('--hide-scrollbars')
    options.add_argument("--disable-infobars")
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-browser-side-navigation')
    # options.add_argument('--disable-extensions')
    options.add_argument('--log-level=0')
    # options.add_argument('--single-process')
    options.add_argument('--ignore-certificate-errors')
    options.add_argument("--disable-plugins-discovery")
    options.add_argument("--start-maximized")
    if use_proxy:
        options.add_argument(
            f"--load-extension={os.path.join(os.path.dirname(os.path.abspath(__file__)), 'plugin', 'zenmate')}")
    driver = uc.Chrome(headless=False, options=options, version_main=version_number)
    rootLogger.debug('Setting up stealth')
    stealth(driver,
            languages=["en-US", "en"],
            vendor="Google Inc.",
            platform="Win32",
            webgl_vendor="Intel Inc.",
            renderer="Intel Iris OpenGL Engine",
            fix_hairline=True,
            )
    for link_ele in read_links(link_list):
        rootLogger.info('Scanning: {}'.format(link_ele))
        try:
            time.sleep(10)
            scrape_obj = None
            for scrape_ele in ScraperBase.__subclasses__():
                if scrape_ele.get_keyword() in link_ele:
                    rootLogger.info('{} website detected'.format(scrape_ele.get_keyword()))
                    scrape_obj = scrape_ele(link_ele, out_path, driver)
                    break
            if scrape_obj is not None:
                scrape_obj.run_browser()
            else:
                raise Exception('Website not supported')
        except Exception as exc:
            rootLogger.error('Error with execution. Omitting link')
            rootLogger.debug('Details: {}'.format(str(exc)))
    driver.close()
    driver.quit()
    rootLogger.info('Goodbye')
